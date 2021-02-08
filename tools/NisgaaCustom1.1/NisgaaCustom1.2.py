#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
# Name:                       Advanced Geospatial Analysis
#                            CUSTOM -- Import--Clip--Append
#                                    version 1.2
#
# Author:      James Burton|Government of British Columbia
#              Geospatial Analyst: Sitkine-Skeena District
#
# Created:     December 2019
#
# Introduction:The primary purpose of this tool is to recieve data from the
#              user, clip the data to the study area, compare shapefile names
#              to a preexisting .xlsx and populate a column based on if the
#              layer is present or absent within the study area, or
#              alternatively if the layer was not included in the analysis.
#              Additional layers included in the analysis but do not meet the
#              criteria headings in the xlsx are appended to the bottom of the
#              spreadsheet. Layers are determined to be present or absent
#              through the calculation of total area after being clipped to
#              study area.
#
#              For a more general application of this tool, please see the
#              GENERAL -- Import -- Clip -- Create xlsx tool which include the
#              framework for additional geometery types. This general purpose
#              tool pairs well with the GENERAL -- Weighted Overlay Script.
#
# Specifics:   In order for this tool to preform optimally, please consider the
#              following items:
#                  1. This tool is for use with FGDB polygon shapefiles.
#                  2. Please ensure shapefiles are named 'Noun_HabitatType_etc'.
#                       For example: "Plants_Freshwater" or "Steelhead_Marine"
#                  3.
#
#              To begin, enter your original GDB file path. Be sure to enter
#              this file path within double or single quotation marks
#              (" " or ' ').
#
#              Second, enter your xlsx template file path. Be sure to enter
#              this file path within double or single quotation marks
#              (" " or ' ').
#
#              The program will, upon successful confirmation of the file paths,
#              begin to execute. The program will create a TEMPORARY FOLDER
#              which will house a TEMPORARY GDB. The delete code is included as
#              a safety measure within the CreateTemp function, and if added to
#              the conclusion of the script, will delete the TEMPORARY items.
#
#              The program will confirm the existance of the template report to
#              copy and populate. Version 2 of this code should feature the
#              creation of a blank xlsx in the event no template exists.
#
#              The program will then copy all the polygon shapefiles from the
#              original GDB to the temp GDB. The study area will be isolated
#              from the polygons. The copied shapefiles are ammended to include
#              a new field in the attribute table to display area
#              (in Hectares, as "POLY_AREA").
#
#              Freshwater salmon species ranges are merged together, with the
#              original, idividual layers being discarded after merger.
#
#              All polygon shapefiles, excluding the study area, are clipped to
#              the extent of the study area.
#
#              The xlsx is then opened and three dictionaries are created.
#              Dictionary one is of column one, feature type.
#              Dictionary two is of column two, feature.
#              Dictionary three is column one and two as a shared value for
#              handling the same features in different feature types.
#              All dictionaries are structured as key:value, with the key
#              representing row number and value representing the column(s).
#
#              The program then generates a list of all the exisiting, clipped
#              polygons. The program then compares the shapefile names to a
#              dictionary to generate the row number in which the feature will
#              be identified as "Present" or "Absent" within the xlsx document.
#              The program will then review the Present/Absent column for
#              'No Data' and insert "N/A" to represent that the feature was not
#              included in the analysis. Layers that may not have an exact
#              match to a feature type & feature are appended to the bottom of
#              the xlsx.
#
# Outputs:     **root\ - derived from template report location**
#              root\xxTemp(today's date) folder
#              root\xxTemp(today's date)_folder\MyTempGDB.gdb
#              root\MyReport(today's date).xlsx
#
# Dependancies: Python 2.7, ArcCatalog 10.x to view temp gdb
#               Must have 2 polygon shapefiles (1 study area, 1 for analysis)
#
# Copyright:   (c) bc.gov.ca, James Burton 2019
#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
# Edit History:
#
# Date: Dec 27, 2019
# Author: James Burton
# Modification Notes: Implemented hard stop at 5 attempts for all user inputs.
# Amalgamated Get Study Area and Check Study Area function, moved below polygon
# extraction. Implemented user input for report name. Removed xlwt module
#
# Date:
# Author:
# Modification Notes:
#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
## Import system modules
import os, arcpy, shutil, openpyxl, time, datetime, xlrd
from arcpy import env
from datetime import datetime, date
from openpyxl import workbook
from itertools import *

## Encode data to utf-8
os.environ["PYTHONIOENCODING"] = "utf-8"

## starts program timer
t0= time.clock()

## sets date for file name
now = datetime.now()
Today= now.strftime('%Y_%m_%d')

## function to obtain original gdb from user
def GetOriginalGDB():

    OriginalGDB = ''
    i = 0
    while i<5:
        try:
            OriginalGDB = input('Enter the GDB file path in quotations (" ")')
            if arcpy.Exists(OriginalGDB):
                break
            else:
                print 'try again'
                i+=1
        except:
            NameError
            print 'That was not entered within "quotation" marks.'
            i+=1
    if i == 5:
        print 'Error, exiting program'
        exit()

    global OriginalGDB
    print 'That file path exists.'
    print ''
    print OriginalGDB

## function to make temporary folder and gdb
def GetTempFolder():

    MyTempContainer=os.path.join(Container+ '\\Temp%s' % Today)
    if os.path.exists(MyTempContainer):
        shutil.rmtree(MyTempContainer)
        os.makedirs(MyTempContainer)

    else:
        os.makedirs(MyTempContainer)

    MyTempGDB = arcpy.CreateFileGDB_management(MyTempContainer, 'MyTempGDB.gdb', '10.0')
    global MyTempContainer
    global MyTempGDB

## function to copy template xlsx document
def CopyXlsxTemplate():
    MyReportIn = ''
    i = 0
    while i<5:
        try:
            MyReportIn = raw_input("Enter your report's name in quotations (' ')")
            break
        except:
            NameError
            print 'That was not entered within "quotation" marks.'
            i+=1
    if i == 5:
        print 'Error, exiting program'
        exit()

    MyReport=os.path.join(Container, MyReportIn + '_'+Today + '.xlsx')
    shutil.copy(TemplateReport, MyReport)
    global MyReport

## function to obtain template xlsx document
def GetXlsxTemplate():

    TemplateReport=''
    i = 0
    while i<5:
        try:
            TemplateReport = input('Enter the template report path in quotations (" ")')
            if os.path.exists(TemplateReport):
                break
            else:
                print 'try again'
                i+=1
        except:
            NameError
            print 'That was not entered within "quotation" marks.'
            i+=1
    if i == 5:
        print 'Error, exiting program'
        exit()


    global TemplateReport
    print 'That file path exists.'
    print ''
    Container=os.path.dirname(TemplateReport)
    global Container
    #CopyXlsxTemplate()

PgFCSList=[]


## function to open xlsx using openpyxl
def OpenXlo():

    print ' openxlo--------------------------------------------------------------------------------------------------------------'
    MyWBo = openpyxl.load_workbook(MyReport)
    global MyWBo
    MyWSo = MyWBo.active
    global MyWSo

## function to open xlsx using xlrd. This is only for xlrd.nrows function later
def OpenXlx():

    print ' openxlx--------------------------------------------------------------------------------------------------------------'
    print 'opening wb and ws'
    MyWBx = xlrd.open_workbook(MyReport, on_demand = True)
    global MyWBx
    MyWSx = MyWBx.sheet_by_index(0)
    global MyWSx

## function to clip polygons based on study area
def ClipPolygons():

    print ' clippolygons--------------------------------------------------------------------------------------------------------------'
    arcpy.env.workspace = str(MyTempGDB)
    PgFCS = arcpy.ListFeatureClasses('', 'polygon')
    for PgFC in PgFCS:

        if PgFC != StudyArea:
            arcpy.Clip_analysis(PgFC, StudyArea, PgFC + '_clip','')

## function to create list of clipped polygons in temp gdb
def MakePgFCSList():

    print ' making PgFCS list--------------------------------------------------------------------------------------------------------------'
    arcpy.env.workspace = str(MyTempGDB)
    PgFCS = arcpy.ListFeatureClasses('', 'polygon')
    for PgFC in PgFCS:
            lookup = 'clip'
            alook= str(PgFC).find(lookup)

            if alook != -1:
                PgList = {}
                PgList = str(PgFC)
                PgFCSList.append(PgList)

            global PgFCSList
    print PgFCSList

## function to extract polygons for original gdb, add area value
def ExtractPolygons():

    print ' extractpolygons--------------------------------------------------------------------------------------------------------------'
    arcpy.env.workspace = OriginalGDB
    PgFCS = arcpy.ListFeatureClasses('', 'polygon')

    for PgFC in PgFCS:
        arcpy.env.workspace = OriginalGDB
        arcpy.FeatureClassToFeatureClass_conversion(PgFC, MyTempGDB,PgFC, '')
        arcpy.env.workspace = str(MyTempGDB)
        arcpy.AddGeometryAttributes_management(PgFC, 'AREA', '', 'HECTARES','')

## function to merge multiple freshwater salmon species polygons
def FreshwaterSalmonMerger():

    print ' joining salmon layers--------------------------------------------------------------------------------------------------------------'
    arcpy.env.workspace=str(MyTempGDB)
    allLayers = arcpy.ListFeatureClasses('','polygon')
    slmList =[]
    i = 0
    sizeofList = len(allLayers)

    ## iterates over all polygons in temp gdb to find only salmon layers that
    ## are NOT marine salmon species, creates a list.
    while i < sizeofList :
        slm_string=str(allLayers[i])
        slm_avoid=slm_string.find('Marine')
        slm_look=slm_string.find('Salmon')

        if slm_avoid == -1:

            if slm_look != -1:
                slmList.append(str(allLayers[i]))

        i += 1

    ## merge individual freshwater salmon layers from salmon list
    if 1< len(slmList) :
        arcpy.Merge_management(slmList, str(MyTempGDB)+ '\\Salmon_FreshwaterNA')
        ii=0
        sizeofslmnlist=len(slmList)

        ## delete individual salmon layers in temp gdb
        while ii < sizeofslmnlist:
            arcpy.env.workspace=str(MyTempGDB)

            if arcpy.Exists(slmList[ii]):
                arcpy.Delete_management(slmList[ii])
            ii+=1

## function to get study area
def GetStudyArea():

    print ' getstudyarea--------------------------------------------------------------------------------------------------------------'
    StudyArea=''
    i = 0
    while i<5:
        try:
            StudyArea = input('Enter the study area name as it appears in the GDB, in quotations (" ")')
            StudyArea2 = str(StudyArea)
            print StudyArea2
            destination = str(MyTempGDB)
            StudyAreaPath = os.path.join(destination,StudyArea2)
            StudyAreaPath2 = str(StudyAreaPath)
            if arcpy.Exists(StudyAreaPath):
                break
            else:
                print 'Try again'
                i+=1
        except:
            NameError
            print 'That was not entered within "quotation" marks.'
            i+=1
    if i == 5:
        print 'Error, exiting program'
        exit()

    global StudyArea

## function to load all xlsx column categories
def LoadXlsxLayers():

    print ' loadxlsxlayers--------------------------------------------------------------------------------------------------------------'
    OpenXlx()
    XlLayersType=[]
    XlLayersList=[]

    ## creates dictionary from column 1 values and corresponding row numbers
    ##(as key)
    for col1 in range(0, MyWSx.nrows):
        XlLayerT = {}
        XlLayerT=MyWSx.cell_value(col1,0)
        XlLayerT2=str(XlLayerT)
        XlLayersType.append(XlLayerT2)
        dict1=dict([(col1, str(MyWSx.cell_value(col1,0))) for col1 in range(0,MyWSx.nrows)])
        global dict1

    ## creates dictionary from column 2 values and corresponding row numbers
    ##(as key)
    for col2 in range(0,MyWSx.nrows):
        XlLayerM = {}
        XlLayerM = unicode(MyWSx.cell_value(col2,1))
        dict2=dict([(col2, unicode(MyWSx.cell_value(col2,1)).encode('utf-8',errors='strict')) for col2 in range(0,MyWSx.nrows)])
        global dict2

    ## creates dictionary joining column 1 and 2 values and matches to the row
    ##number, as a key
    dict3 = {key: value + ', ' + dict2[key] for key, value in dict1.iteritems()}
    global dict3

## custom function to find polygon layer in temp gdb, ascertain if the layer has
##area inside of the study area, populates corresponding xlsx row with "present"
##or "absent" values accordingly
def RetrieveAreaPg(workspace,PgName,RowToPopulate):

    print ' retrieveareapg--------------------------------------------------------------------------------------------------------------'
    #print RowToPopulate
    arcpy.env.workspace = str(workspace)
    PgFCSa = arcpy.SearchCursor(str(PgName))
    PgFCSaList=[]
    for PgFCa in PgFCSa:
        TP1CArea=PgFCa.getValue('POLY_Area')
        PgaList={}
        PgaList=TP1CArea
        PgFCSaList.append(PgaList)

    xarea= sum(PgFCSaList)
    if xarea is not None:
        OpenXlo()
        mycell= MyWSo.cell(row=RowToPopulate, column=3)
        mycell.value='Present'
        MyWBo.save(MyReport)

    else:
        OpenXlo()
        mycell= MyWSo.cell(row=RowToPopulate, column=3)
        mycell.value='Absent'
        MyWBo.save(MyReport)

## function to handle layers not matched to dictionary
## appends layer name to bottom of xlsx
def MakeXLSXTitle(LayerName,RowToPopulate):

    print ' MakeXlSXTitle--------------------------------------------------------------------------------------------------------------'
    OpenXlo()
    mycell= MyWSo.cell(row=RowToPopulate, column=2)
    mycell.value=LayerName
    MyWBo.save(MyReport)

## function to handle single dictionary searches
## used for layers with no multiple occurances between habitat types
def SearchLayers(dictionary, layercode):

    SearchLayers.a=[]
    for key, value in dictionary.items():

        if layercode in value:
            SearchLayers.a.append(key)

## function to handle multiple dictionary searches
## used for layers with multiple occurances between habitat types, for example
def SearchLayers2(dictionary, layercode, referencestr):

    SearchLayers2.b=[]
    for key, value in dictionary.items():

        if layercode in value:

            if referencestr in value:
                SearchLayers2.b.append(key)

## function to match layers to column values
## function contains specific elif's for client
def GetPgFCSList():

    MakePgFCSList()
    ai = 0
    sizeofList = len(PgFCSList)
    while ai < sizeofList :
        a_string=str(PgFCSList[ai])

        lookup1 = 'Bear'
        lookup2 = 'Huckle'
        lookup3 = 'Freshwater'
        lookup4 = 'Marine'
        lookup5 = 'UWR'
        lookup6 = 'SRB'
        lookup7 = 'RSPF'
        lookup8 = 'hydro'
        lookup9 = 'Mushroom'

        if a_string[:5] == 'Black':
            alook1= a_string.find(lookup1)

            if alook1 != -1:
                SearchLayers2(dict3, a_string[:5],lookup1)
                Position = int(SearchLayers2.b[0])+1
                RetrieveAreaPg(MyTempGDB,a_string,Position)

            else:
                SearchLayers2(dict3, a_string[:5],lookup2)
                Position = int(SearchLayers2.b[0])+1
                RetrieveAreaPg(MyTempGDB,a_string,Position)

        elif a_string[:5] =='Fish_':
            alook3=a_string.find(lookup3)

            if alook3 != -1:
                SearchLayers2(dict3, 'Non-salmon', lookup3)
                Position = int(SearchLayers2.b[0])+1
                RetrieveAreaPg(MyTempGDB,a_string,Position)

            else:
                SearchLayers2(dict3, 'Non-salmon', lookup4)
                Position = int(SearchLayers2.b[0])+1
                RetrieveAreaPg(MyTempGDB,a_string,Position)

        elif a_string[:5]=='Plant':
            alook4=a_string.find(lookup3)

            if alook4 != -1:
                SearchLayers2(dict2, str.lower(a_string[:5]), lookup3)
                Position = int(SearchLayers2.b[0])+1
                RetrieveAreaPg(MyTempGDB,a_string,Position)

            else:
                SearchLayers2(dict2, a_string[:5], lookup4)
                Position = int(SearchLayers2.b[0])+1
                RetrieveAreaPg(MyTempGDB,a_string,Position)

        elif a_string[:5] == 'Steel' or a_string[:5] =='Salmo':
            alook4=a_string.find(lookup3)

            if alook4 != -1:
                SearchLayers2(dict3, a_string[:5], lookup3)
                Position = int(SearchLayers2.b[0])+1
                RetrieveAreaPg(MyTempGDB,a_string,Position)

            else:
                SearchLayers2(dict3, a_string[:5], lookup4)
                Position = int(SearchLayers2.b[0])+1
                RetrieveAreaPg(MyTempGDB,a_string,Position)

        elif a_string[:5] == 'Moose' or a_string[:5] == 'MtnGo':
            alook5=a_string.find(lookup5)
            alook6=a_string.find(lookup6)
            alook7=a_string.find(lookup7)

            if alook5 != -1 and a_string[:5]=='Moose':
                SearchLayers2(dict3, a_string[:5], lookup5)
                Position = int(SearchLayers2.b[0])+1
                RetrieveAreaPg(MyTempGDB,a_string,Position)

            elif alook6 != -1:
                SearchLayers2(dict3, a_string[:5], lookup6)
                Position = int(SearchLayers2.b[0])+1
                RetrieveAreaPg(MyTempGDB,a_string,Position)

            elif alook7 !=-1:
                SearchLayers2(dict3, 'Goat', lookup7)
                Position = int(SearchLayers2.b[0])+1
                RetrieveAreaPg(MyTempGDB,a_string,Position)

            else:
                SearchLayers2(dict3, 'Goat', lookup5)
                Position = int(SearchLayers2.b[0])+1
                RetrieveAreaPg(MyTempGDB,a_string,Position)

        elif a_string[:5]=='Hydro':
            SearchLayers(dict3, lookup8)
            Position = int(SearchLayers.a[0])+1
            RetrieveAreaPg(MyTempGDB,a_string,Position)

        elif a_string[:5]=='PineM':
            SearchLayers(dict3, lookup9)
            Position = int(SearchLayers.a[0])+1
            RetrieveAreaPg(MyTempGDB,a_string,Position)

        elif a_string[:5] =='Herit':
            SearchLayers(dict2, 'Heritage')
            Position = int(SearchLayers.a[0])+1
            RetrieveAreaPg(MyTempGDB,a_string,Position)

        elif a_string[:5] == 'Sooty':
            SearchLayers(dict2, 'Grouse')
            Position = int(SearchLayers.a[0])+1
            RetrieveAreaPg(MyTempGDB,a_string,Position)

        elif a_string[:5] == 'Nisga':
            OpenXlo()
            row_count = (MyWSo.max_row)+1
            MakeXLSXTitle(a_string,row_count)
            RetrieveAreaPg(MyTempGDB,a_string,row_count)

        else:
            SearchLayers(dict3, a_string[:5])

            try:
                Position = int(SearchLayers.a[0])+1
                RetrieveAreaPg(MyTempGDB,a_string,Position)

            except:
                IndexError
                OpenXlo()
                row_count = (MyWSo.max_row)+1
                MakeXLSXTitle(a_string,row_count)
                RetrieveAreaPg(MyTempGDB,a_string,row_count)
        ai += 1

## function to handle blank cells in present/absent column
## if cell contains no data, N/A will be populated in the cell value
def BlankCheck():

    print ' Blank Check--------------------------------------------------------------------------------------------------------------'
    OpenXlo()
    row_count = (MyWSo.max_row)
    i=1
    while i < row_count:

        if MyWSo.cell(row = i, column=3).value == None:
            mycell= MyWSo.cell(row=i, column=3)
            mycell.value='N/A'
            MyWBo.save(MyReport)
        i+=1


GetOriginalGDB()
GetXlsxTemplate()
GetTempFolder()
CopyXlsxTemplate()
ExtractPolygons()
GetStudyArea()
FreshwaterSalmonMerger()
ClipPolygons()
LoadXlsxLayers()
GetPgFCSList()
BlankCheck()
MyWBo.save(MyReport)

print 'DONE :)'
t1 = time.clock() - t0
print('Time elapsed: ', (t1 - t0)) # CPU seconds elapsed (floating point)